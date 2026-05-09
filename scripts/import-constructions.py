"""Import V11 Construction Bank from Excel to JSON/TS for DB seeding"""
import openpyxl, json, sys

V11_PATH = r'f:\桌面\构式二代\constructions-main\V11.xlsx'
OUT_JSON = r'f:\桌面\构式二代\constructions-main\src\data\v11-constructions.json'

wb = openpyxl.load_workbook(V11_PATH)

level_map = {
    '01_初中版构式库': 'junior',
    '02_高中版构式库': 'senior',
    '03_雅思基础版构式库': 'ielts_basic',
    '04_雅思进阶版构式库': 'ielts_advanced',
}

all_constructions = []

for sn in level_map:
    ws = wb[sn]
    level = level_map[sn]

    # Find header row
    hrow = None
    for r in ws.iter_rows(min_row=1, max_row=20):
        for c in r:
            if c.value and '构式编号' in str(c.value):
                hrow = c.row
                break
        if hrow:
            break
    if not hrow:
        print(f'  WARN: no header in {sn}')
        continue

    # Read all rows
    count = 0
    current_category = ''
    for row in ws.iter_rows(min_row=hrow + 1, max_row=ws.max_row, values_only=True):
        code = str(row[0]).strip() if row[0] else ''
        if not code:
            continue

        # Category rows like "【S1_动词短语构式】"
        if code.startswith('【') and code.endswith('】'):
            current_category = code.strip('【】')
            continue

        name = str(row[1]).strip() if len(row) > 1 and row[1] else ''
        template = str(row[2]).strip() if len(row) > 2 and row[2] else ''
        core_words = str(row[3]).strip() if len(row) > 3 and row[3] else ''
        function = str(row[4]).strip() if len(row) > 4 and row[4] else ''
        usage_note = str(row[5]).strip() if len(row) > 5 and row[5] else ''
        example = str(row[6]).strip() if len(row) > 6 and row[6] else ''
        variants = str(row[7]).strip() if len(row) > 7 and row[7] else ''
        difficulty = str(row[8]).strip() if len(row) > 8 and row[8] else ''

        all_constructions.append({
            'code': code,
            'name': name,
            'template': template,
            'coreWords': core_words,
            'function': function,
            'usageNote': usage_note,
            'example': example,
            'variants': variants if variants else None,
            'difficulty': difficulty,
            'level': level,
            'category': current_category,
        })
        count += 1

    print(f'  {sn}: {count} entries')

print(f'\nTotal: {len(all_constructions)} constructions')

# Save JSON
with open(OUT_JSON, 'w', encoding='utf-8') as f:
    json.dump(all_constructions, f, ensure_ascii=False)
print(f'Saved to {OUT_JSON}')
